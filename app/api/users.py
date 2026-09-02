"""User management API endpoints — CRUD, batch import, profile."""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import select, func, delete
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.security import get_admin_user, get_current_user, hash_password
from app.models import User as UserModel, UserSubject, Subject
from app.schemas.schemas import (
    ImportResult,
    PaginatedResponse,
    UserCreate,
    UserOut,
    UserUpdate,
)


class SubjectIdsBody(BaseModel):
    subject_ids: list[int]


router = APIRouter(prefix="/api/users", tags=["users"])


@router.get("", response_model=PaginatedResponse)
async def list_users(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    role: str | None = None,
    class_name: str | None = None,
    search: str | None = None,
    subject_id: int | None = None,
    online_status: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """List users with optional filters. Teachers see their class only."""
    query = select(UserModel)

    # Role-based visibility
    if current_user["role"] == "teacher":
        query = query.where(UserModel.role == "student")
    elif current_user["role"] == "student":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权查看用户列表")

    if role:
        query = query.where(UserModel.role == role)
    if class_name:
        query = query.where(UserModel.class_name == class_name)
    if search:
        query = query.where(
            (UserModel.username.ilike(f"%{search}%"))
            | (UserModel.real_name.ilike(f"%{search}%"))
            | (UserModel.student_id.ilike(f"%{search}%"))
        )
    if subject_id:
        subq = select(UserSubject.user_id).where(UserSubject.subject_id == subject_id)
        query = query.where(UserModel.id.in_(subq))

    # Count total
    count_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_query)).scalar() or 0

    # Fetch page
    offset = (page - 1) * page_size
    query = query.order_by(UserModel.created_at.desc()).offset(offset).limit(page_size)
    result = await db.execute(query)
    users = result.scalars().all()

    # Batch-load all subjects for these users (avoid N+1)
    user_ids = [u.id for u in users]
    subject_map: dict[int, list[dict]] = {uid: [] for uid in user_ids}
    if user_ids:
        from sqlalchemy import select as sa_select
        subj_result = await db.execute(
            sa_select(UserSubject.user_id, Subject.id, Subject.name)
            .join(Subject, Subject.id == UserSubject.subject_id)
            .where(UserSubject.user_id.in_(user_ids))
        )
        for user_id, subj_id, subj_name in subj_result.all():
            subject_map.setdefault(user_id, []).append({"id": subj_id, "name": subj_name})

    # Enrich with subjects
    user_list = []
    for u in users:
        u_dict = UserOut.model_validate(u).model_dump()
        u_dict["subjects"] = subject_map.get(u.id, [])
        user_list.append(u_dict)

    return PaginatedResponse(
        items=user_list,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=max(1, (total + page_size - 1) // page_size),
    )


class UserCreateBody(UserCreate):
    subject_ids: list[int] = []


@router.post("", response_model=UserOut, status_code=status.HTTP_201_CREATED)
async def create_user(
    data: UserCreateBody,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_admin_user),
):
    """手动创建单个用户（管理员）。"""
    existing = await db.execute(select(UserModel).where(UserModel.username == data.username))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="用户名已存在")

    user = UserModel(
        username=data.username,
        password_hash=hash_password(data.password),
        real_name=data.real_name,
        role=data.role,
        class_name=data.class_name,
        student_id=data.student_id,
        email=data.email,
    )
    db.add(user)
    await db.flush()  # 拿到 user.id

    for sid in data.subject_ids:
        db.add(UserSubject(user_id=user.id, subject_id=sid))

    await db.commit()
    await db.refresh(user)
    return UserOut.model_validate(user)


@router.get("/classes")
async def list_classes(
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """List distinct class names for filtering."""
    result = await db.execute(
        select(UserModel.class_name).where(UserModel.class_name.isnot(None)).distinct()
    )
    classes = [r[0] for r in result.all() if r[0]]
    classes.sort()
    return classes


@router.get("/{user_id}", response_model=UserOut)
async def get_user(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get a single user by ID."""
    if current_user["role"] == "student" and current_user["user_id"] != user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权查看其他用户信息")

    result = await db.execute(select(UserModel).where(UserModel.id == user_id))
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    return UserOut.model_validate(user)


@router.put("/{user_id}", response_model=UserOut)
async def update_user(
    user_id: int,
    data: UserUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Update user info. Admin-only for role/status changes."""
    result = await db.execute(select(UserModel).where(UserModel.id == user_id))
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")

    # Permission check
    if current_user["role"] != "admin" and current_user["user_id"] != user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="无权修改此用户")
    if current_user["role"] != "admin" and (data.role is not None or data.is_active is not None):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="只有管理员可以修改角色或状态")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(user, key, value)

    await db.flush()
    await db.refresh(user)
    return UserOut.model_validate(user)


@router.delete("/{user_id}")
async def delete_user(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_admin_user),
):
    """Delete a user (admin only)."""
    result = await db.execute(select(UserModel).where(UserModel.id == user_id))
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")
    if user.id == current_user["user_id"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="不能删除自己")

    await db.delete(user)
    return {"message": f"已删除用户: {user.username}"}


@router.get("/{user_id}/subjects")
async def get_user_subjects(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get subjects assigned to a user."""
    result = await db.execute(
        select(Subject).join(UserSubject).where(UserSubject.user_id == user_id)
    )
    return [{"id": s.id, "name": s.name} for s in result.scalars().all()]


@router.put("/{user_id}/subjects")
async def update_user_subjects(
    user_id: int,
    data: "SubjectIdsBody",
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_admin_user),
):
    """Update subjects assigned to a user (admin only). Body: {"subject_ids": [1,2,3]}"""
    # Remove existing
    await db.execute(delete(UserSubject).where(UserSubject.user_id == user_id))
    # Add new
    for sid in data.subject_ids:
        db.add(UserSubject(user_id=user_id, subject_id=sid))
    await db.flush()
    return {"message": "学科分配已更新", "subject_ids": data.subject_ids}


@router.post("/{user_id}/reset-password")
async def reset_user_password(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_admin_user),
):
    """Reset a user's password to their student_id or default."""
    result = await db.execute(select(UserModel).where(UserModel.id == user_id))
    user = result.scalar_one_or_none()
    if user is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="用户不存在")

    new_password = user.student_id or "123456"
    user.password_hash = hash_password(new_password)
    await db.flush()
    return {"message": f"密码已重置为: {new_password}"}


@router.post("/import-excel", response_model=ImportResult)
async def import_users_excel(
    file: UploadFile,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_admin_user),
):
    """Batch import users from Excel (admin only).
    Expected columns: 用户名 | 真实姓名 | 角色 | 班级 | 学号 | 密码(可选)
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅支持 .xlsx / .xls 文件")

    contents = await file.read()
    wb = load_workbook(contents, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    total = len(rows)
    success = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        try:
            username = str(row[0]).strip()
            real_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            role = str(row[2]).strip() if len(row) > 2 and row[2] else "student"
            class_name = str(row[3]).strip() if len(row) > 3 and row[3] else None
            student_id = str(row[4]).strip() if len(row) > 4 and row[4] else None
            password = str(row[5]).strip() if len(row) > 5 and row[5] else "123456"

            if role not in ("student", "teacher", "admin"):
                errors.append(f"行{i}: 无效角色'{role}'")
                continue

            existing = await db.execute(select(UserModel).where(UserModel.username == username))
            if existing.scalar_one_or_none():
                errors.append(f"行{i}: 用户名'{username}'已存在")
                continue

            user = UserModel(
                username=username,
                password_hash=hash_password(password),
                real_name=real_name,
                role=role,
                class_name=class_name,
                student_id=student_id,
            )
            db.add(user)
            success += 1
        except Exception as e:
            errors.append(f"行{i}: {str(e)}")

    await db.flush()
    return ImportResult(total=total, success=success, failed=total - success, errors=errors)
