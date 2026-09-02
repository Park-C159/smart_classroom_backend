"""File upload API — user import, exercise upload."""
import io
import logging

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl

from app.core.database import get_db
from app.core.security import get_current_user, get_teacher_or_admin, get_admin_user, hash_password
from app.models import User, TestQuestion

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/upload", tags=["文件上传"])


_TYPE_ALIASES = {
    "选择": "choice", "选择题": "choice", "choice": "choice",
    "填空": "fill", "填空题": "fill", "fill": "fill",
    "简答": "short_answer", "简答题": "short_answer", "证明": "short_answer",
    "证明题": "short_answer", "short_answer": "short_answer", "解答": "short_answer",
}


def _find_col(header_map: dict, *keys: str):
    for h, idx in header_map.items():
        if not h:
            continue
        hl = str(h).strip()
        for k in keys:
            if k in hl or hl in k:
                return idx
    return None


@router.post("/import-exercises")
async def import_exercises_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_teacher_or_admin),
):
    """批量导入试题到独立试题库（教师/管理员）。

    模板列：题型 | 章节 | 题目 | 选项A | 选项B | 选项C | 选项D | 答案 | 难度
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 文件")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"total": 0, "created": 0, "errors": []}

    headers = [c for c in rows[0]]
    header_map = {h: i for i, h in enumerate(headers)}

    created = 0
    errors = []

    for i, row in enumerate(rows[1:], start=2):
        def cell(*keys):
            idx = _find_col(header_map, *keys)
            return row[idx] if idx is not None and idx < len(row) else None

        try:
            qtype_raw = str(cell("题型", "类型", "type") or "").strip()
            qtype = _TYPE_ALIASES.get(qtype_raw) or _TYPE_ALIASES.get(qtype_raw.lower())
            if not qtype:
                errors.append(f"第{i}行: 无法识别题型'{qtype_raw}'")
                continue

            question_text = str(cell("题目", "题干", "question") or "").strip()
            answer_text = str(cell("答案", "参考答案", "answer") or "").strip()
            if not question_text or not answer_text:
                errors.append(f"第{i}行: 题目或答案为空")
                continue

            options = None
            if qtype == "choice":
                opts = []
                for key in ("A", "B", "C", "D"):
                    v = cell(f"选项{key}", f"选项 {key}", key)
                    if v is not None and str(v).strip():
                        opts.append({"key": key, "text": str(v).strip()})
                if not opts:
                    errors.append(f"第{i}行: 选择题缺少选项")
                    continue
                options = opts

            difficulty = 3
            d_raw = cell("难度", "difficulty")
            if d_raw is not None and str(d_raw).strip():
                try:
                    difficulty = max(1, min(5, int(float(str(d_raw).strip()))))
                except ValueError:
                    pass

            q = TestQuestion(
                subject_id=None,
                chapter=str(cell("章节", "章", "chapter") or "").strip() or None,
                kp_id=None,
                question_type=qtype,
                question_text=question_text,
                options=options,
                answer_text=answer_text,
                difficulty=difficulty,
                created_by=current_user["user_id"],
                verified=False,
            )
            db.add(q)
            created += 1
        except Exception as e:
            errors.append(f"第{i}行: {e}")

    await db.commit()
    return {"total": len(rows) - 1, "created": created, "errors": errors}


@router.post("/import-users")
async def import_users_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_admin_user),
):
    """Batch import users from Excel file."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(400, "仅支持 .xlsx / .xls / .csv 文件")

    content = await file.read()

    # Handle CSV
    if file.filename.endswith(".csv"):
        import csv
        rows = list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

    created = 0
    errors = []

    for i, row in enumerate(rows):
        try:
            name_col = None
            for h in row.keys():
                if h and ("姓名" in str(h) or "name" in str(h).lower()):
                    name_col = h
                    break
            id_col = None
            for h in row.keys():
                if h and ("学号" in str(h) or "student_id" in str(h) or "id" in str(h).lower()):
                    id_col = h
                    break

            real_name = str(row.get(name_col, "")).strip() if name_col else ""
            student_id = str(row.get(id_col, "")).strip() if id_col else ""

            if not real_name:
                continue

            username = student_id or f"stu_{real_name}"
            existing = (await db.execute(select(User).where(User.username == username))).scalar_one_or_none()
            if existing:
                continue

            user = User(
                username=username,
                password_hash=hash_password(student_id or "123456"),
                real_name=real_name,
                role="student",
                student_id=student_id,
                class_name=str(row.get("班级", "")).strip() or None,
            )
            db.add(user)
            created += 1
        except Exception as e:
            errors.append(f"第{i + 2}行: {e}")

    await db.commit()
    return {"created": created, "errors": errors, "total": len(rows)}
